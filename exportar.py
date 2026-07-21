# exportar.py
import pandas as pd
from io import BytesIO
from database import SessionLocal, PosicaoDiaria
from openpyxl.styles import Font, Alignment, Border, Side

def formatar_br(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

ITENS_ORDEM = [
    'CARTEIRA', 'MERCADO PAGO', 'VEICULO', 'SEGURADORA', 'GARANTIA', 'BANCOS',
    'CARTOES', 'NOVOS PAGOS', 'USADOS PAGOS', 'FUNDAO NOVOS', 'FIDIC',
    'H.B.PECAS', 'ESTOQUE PECAS', 'OBRIG. A PAGA', 'ADIANTAMENTOS',
    'TRANSITORIA', 'DIF_TRANS_ADIANT'
]

def gerar_excel_dashboard(data_ref, empresas):
    db = SessionLocal()
    dados = db.query(PosicaoDiaria).filter(
        PosicaoDiaria.data == data_ref,
        PosicaoDiaria.empresa.in_(empresas)
    ).all()
    db.close()

    if not dados:
        return None

    df_pivot = pd.DataFrame(0.0, index=ITENS_ORDEM, columns=empresas)
    df_qtd = pd.DataFrame(0, index=ITENS_ORDEM, columns=empresas)

    for d in dados:
        if d.tipo_titulo in ITENS_ORDEM and d.empresa in empresas:
            df_pivot.loc[d.tipo_titulo, d.empresa] = d.valor
            df_qtd.loc[d.tipo_titulo, d.empresa] = d.qtd_veiculos

    output = BytesIO()
    nome_aba = 'POSIÇÃO ' + data_ref.strftime('%d-%m-%Y')

    df_dummy = pd.DataFrame()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_dummy.to_excel(writer, sheet_name=nome_aba, index=False)
        workbook = writer.book
        worksheet = writer.sheets[nome_aba]

        # Estilos
        bold_font = Font(bold=True, size=11)
        normal_font = Font(size=11)
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Titulo e Data
        worksheet.merge_cells('A1:F1')
        cell_titulo = worksheet.cell(row=1, column=1, value="POSIÇÃO FINANCEIRA DIÁRIA")
        cell_titulo.font = bold_font
        cell_titulo.alignment = center_align

        worksheet.cell(row=1, column=7, value="DATA").font = bold_font
        worksheet.cell(row=1, column=8, value=data_ref.strftime('%d/%m/%Y'))

        row = 2
        col_atual = 1

        for emp in empresas:
            # Cabeçalho Empresa
            worksheet.merge_cells(start_row=row+1, start_column=col_atual, end_row=row+1, end_column=col_atual+1)
            cell_emp = worksheet.cell(row=row+1, column=col_atual, value=emp)
            cell_emp.font = bold_font
            cell_emp.alignment = center_align

            # DESCRICAO | VALORES
            worksheet.cell(row=row+2, column=col_atual, value="DESCRICAO").font = bold_font
            worksheet.cell(row=row+2, column=col_atual+1, value="VALORES").font = bold_font

            linha_item = row + 3
            total_emp = 0
            for item in ITENS_ORDEM:
                valor = df_pivot.loc[item, emp]
                qtd = df_qtd.loc[item, emp]

                if item in ['NOVOS PAGOS', 'USADOS PAGOS']:
                    valor_txt = f"{formatar_br(valor).replace('R$ ', '')} - {qtd}"
                else:
                    valor_txt = formatar_br(valor)

                cell_desc = worksheet.cell(row=linha_item, column=col_atual, value=item)
                cell_val = worksheet.cell(row=linha_item, column=col_atual+1, value=valor_txt)

                cell_desc.font = normal_font
                cell_val.font = normal_font
                cell_val.alignment = right_align

                # Borda
                cell_desc.border = thin_border
                cell_val.border = thin_border

                total_emp += valor
                linha_item += 1

            # TOTAL
            cell_total_desc = worksheet.cell(row=linha_item, column=col_atual, value="TOTAL")
            cell_total_val = worksheet.cell(row=linha_item, column=col_atual+1, value=formatar_br(total_emp))
            cell_total_desc.font = bold_font
            cell_total_val.font = bold_font
            cell_total_val.alignment = right_align
            cell_total_desc.border = thin_border
            cell_total_val.border = thin_border

            col_atual += 3

        # Ajustar largura
        for col in ['A', 'B', 'D', 'E', 'G', 'H']:
            worksheet.column_dimensions[col].width = 18

    return output.getvalue()