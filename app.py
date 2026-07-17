import pandas as pd
import streamlit as st
from datetime import date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from database import SessionLocal, PosicaoDiaria, Usuarios
import hashlib
import random
import string # <- adiciona esses 2
from auth import verificar_login, tela_cadastro_usuario
import numpy as np

def gerar_hash(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def gerar_senha_aleatoria(tamanho=8):
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(tamanho))
    
# Empresas
EMPRESAS = ["MATRIZ", "WS", "EUSEBIO"]

st.set_page_config(layout="wide", page_title="Posição Financeira Diária")
def tela_login():
    st.title("🔒 Acesso Restrito - Posição Diária")
    
    with st.form("login"):
        email = st.text_input("Email")
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar")
        
        if entrar:
            db = SessionLocal()
            senha_hash = hashlib.sha256(senha.encode()).hexdigest()
            user = db.query(Usuarios).filter_by(email=email, senha_hash=senha_hash, ativo=True).first()
            db.close()
            
            if user:
                st.session_state['logado'] = True
                st.session_state['usuario'] = user.nome
                st.session_state['email'] = user.email
                st.session_state['perfil'] = user.perfil # <- ADICIONA ESSA
                st.rerun()
            else:
                st.error("Email ou senha inválidos")

# TRAVA
if 'logado' not in st.session_state:
    st.session_state['logado'] = False

if not st.session_state['logado']:
    tela_login()
    st.stop() # para tudo aqui se não logou

# SIDEBAR
if st.session_state['logado']:
    # Pega o perfil do usuário do banco SÓ se estiver logado
    db = SessionLocal()
    user_db = db.query(Usuarios).filter_by(email=st.session_state['email']).first()
    st.session_state['perfil'] = user_db.perfil if user_db else "Usuario"
    db.close()
    
    st.sidebar.write(f"Perfil: {st.session_state['perfil']}")

    # <- COLA AQUI: Só aparece se for Admin 
    if st.session_state['perfil'] == 'Admin':
        with st.sidebar.expander("👥 Gerenciar Usuários"):
            # 1. LISTA DE USUÁRIOS
            st.markdown("#### Usuários Cadastrados")
            db = SessionLocal()
            users = db.query(Usuarios).order_by(Usuarios.id.desc()).all()
            db.close()
    
            selected_id = None
            if users:
                df_users = pd.DataFrame([{
                    'ID': u.id,
                    'Nome': u.nome,
                    'Email': u.email,
                    'Perfil': u.perfil,
                    'Ativo': 'Sim' if u.ativo else 'Não'
                } for u in users])
    
                selected = st.dataframe(
                    df_users,
                    use_container_width=True,
                    hide_index=True,
                    height=200,
                    on_select="rerun",
                    selection_mode="single-row"
                )
    
                if selected['selection']['rows']:
                    idx = selected['selection']['rows'][0]
                    selected_id = int(df_users.iloc[idx]['ID']) # <- CORRIGE ERRO DO POSTGRES
            else:
                st.warning("Nenhum usuário cadastrado")
    
            st.divider()
    
            # 2. FORM DE EDIÇÃO / CADASTRO
            st.markdown("#### Dados do Usuário")
    
            user_edit = None
            if selected_id:
                db = SessionLocal()
                user_edit = db.query(Usuarios).filter(Usuarios.id == selected_id).first()
                db.close()
    
            with st.form("form_usuario", clear_on_submit=False):
                nome = st.text_input("Nome", value=user_edit.nome if user_edit else "")
                email = st.text_input("Email", value=user_edit.email if user_edit else "", disabled=bool(user_edit))
                perfil = st.selectbox("Perfil", ["Usuario", "Admin"], index=0 if not user_edit or user_edit.perfil=="Usuario" else 1)
                ativo = st.checkbox("Ativo", value=user_edit.ativo if user_edit else True)
    
                col1, col2, col3 = st.columns(3)
                with col1:
                    btn_salvar = st.form_submit_button("💾 Salvar", use_container_width=True)
                with col2:
                    btn_senha = st.form_submit_button("🔑 Nova Senha", use_container_width=True)
                with col3:
                    # LÓGICA DO BOTÃO: muda se estiver ativo ou não
                    if user_edit:
                        if user_edit.ativo:
                            btn_toggle = st.form_submit_button("🗑️ Desativar", use_container_width=True, type="primary")
                        else:
                            btn_toggle = st.form_submit_button("✅ Ativar", use_container_width=True, type="secondary")
                    else:
                        btn_toggle = st.form_submit_button("🗑️ Desativar", use_container_width=True, disabled=True)
    
            # 3. LÓGICA DOS BOTÕES
            db = SessionLocal()
    
            # CADASTRAR OU EDITAR
            if btn_salvar:
                if user_edit: # EDITAR
                    user_edit.nome = nome
                    user_edit.perfil = perfil
                    user_edit.ativo = ativo
                    msg = "Usuário atualizado!"
                else: # CADASTRAR
                    if db.query(Usuarios).filter_by(email=email).first():
                        st.error("Erro: Este email já está cadastrado")
                        db.close()
                        st.stop()
                    senha_temp = gerar_senha_aleatoria() # <- SENHA ALEATÓRIA
                    novo = Usuarios(nome=nome, email=email, senha_hash=gerar_hash(senha_temp), perfil=perfil, ativo=ativo)
                    db.add(novo)
                    msg = f"Usuário cadastrado! Senha padrão: `{senha_temp}`"
    
                db.commit()
                st.success(msg)
                db.close()
                st.rerun()
    
            # TROCAR SENHA
            if btn_senha and user_edit:
                nova_senha = gerar_senha_aleatoria() # <- CORRIGE NameError
                user_edit.senha_hash = gerar_hash(nova_senha)
                db.commit()
                st.success(f"Senha resetada! Nova senha: `{nova_senha}`")
                db.close()
                st.rerun()
    
            # ATIVAR / DESATIVAR
            if btn_toggle and user_edit:
                user_edit.ativo = not user_edit.ativo
                status = "Ativado" if user_edit.ativo else "Desativado"
                db.commit()
                st.warning(f"Usuário {status}")
                db.close()
                st.rerun()
    
            db.close()
    
            # 4. CADASTRO RÁPIDO - se não selecionou ninguém
            if not selected_id:
                st.divider()
                st.markdown("#### Cadastrar Novo")
                tela_cadastro_usuario() # sua função antiga  
                
st.title("Dashboard Financeira Diária")
st.markdown("""
<style>
[data-testid="stMetricValue"] {
    font-size: 20px!important;
    display: flex!important;
    align-items: baseline!important;
    gap: 3px!important;
}
[data-testid="stMetricValue"]::before {
    content: "R$ ";
    font-size: 12px!important;
    font-weight: 600;
}
[data-testid="stMetricLabel"] {
    font-size: 11px!important;
}
</style>
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["Lançamento", "Histórico"])

def formatar_br(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
# Converter Valor
def converter_valor_br(valor):
    if pd.isna(valor): return 0.0
    val = str(valor).strip()
    if val == '' or val == '-' or val.upper() == 'NAN': return 0.0
    val = val.replace('R$', '').replace(' ', '')
    if ',' in val: val = val.replace('.', '').replace(',', '.')
    try: return float(val)
    except: return 0.0

def detectar_empresa(nome):
    nome = str(nome).upper()
    if 'MATRIZ' in nome: return 'MATRIZ'
    if 'WS' in nome: return 'WS'
    if 'EUSEBIO' in nome: return 'EUSEBIO'
    return 'OUTROS'

@st.cache_data
def get_total_empresa(data, empresa):
    """Calcula o total líquido da empresa na data"""
    # Garante que data é objeto date
    if isinstance(data, str):
        data = date.fromisoformat(data)
        
    db = SessionLocal()
    regs = db.query(PosicaoDiaria).filter(
        PosicaoDiaria.data == data,
        PosicaoDiaria.empresa == empresa
    ).all()
    db.close()
    if not regs: return 0.0

    total = 0.0
    for r in regs:
        if r.tipo_titulo == 'OBRIG. A PAGA':
            total -= r.valor
        elif r.tipo_titulo not in ['TRANSITORIA', 'DIF_TRANS_ADIANT']:
            total += r.valor
    return total

@st.cache_data
def get_variacao_empresa(data_hoje, data_ontem, empresa):
    """Calcula a variação % vs dia anterior"""
    # Converte pra date pra não bugar o cache
    if isinstance(data_hoje, str): data_hoje = date.fromisoformat(data_hoje)
    if isinstance(data_ontem, str): data_ontem = date.fromisoformat(data_ontem)
        
    total_hoje = get_total_empresa(data_hoje, empresa)
    total_ontem = get_total_empresa(data_ontem, empresa)

    if total_ontem == 0:
        return None

    variacao = ((total_hoje - total_ontem) / abs(total_ontem)) * 100
    return variacao

def formatar_compacto(valor):
    """Formata para caber no card: 1.2M, 589K"""
    if valor >= 1_000_000:
        return f"{valor/1_000_000:.2f}M"
    elif valor >= 1_000:
        return f"{valor/1_000:.1f}K"
    else:
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def normalizar_tipo(titulo):
    t = str(titulo).upper().strip()
    if 'CARTEIRA' in t: return 'CARTEIRA'
    if 'MERCADO.PAGO' in t or 'MERCADO PAGO' in t: return 'MERCADO PAGO'
    if 'VEICULO' in t or 'V-VEICULO' in t or 'VENDA VEIC' in t: return 'VEICULO'
    if 'SEGURADORA' in t or 'SEGURO' in t: return 'SEGURADORA'
    if 'GARANTIA' in t or 'VENDA GARANTIA' in t: return 'GARANTIA'
    if 'BANCO' in t: return 'BANCOS'
    if 'CARTAO' in t or 'CARTÕES' in t or 'CARTAO DE CREDITO' in t or 'CREDITO' in t: return 'CARTOES'
    if 'FUNDAO NOVOS' in t or 'FUNDAO' in t: return 'FUNDAO NOVOS'
    if 'NOVOS PAGOS' in t or 'NOVOS.PAGOS' in t: return 'NOVOS PAGOS'
    if 'USADOS PAGOS' in t or 'USADOS.PAGOS' in t or 'USADOS' in t: return 'USADOS PAGOS'
    if 'FIDIC' in t: return 'FIDIC'
    if 'H.B.PECAS' in t or 'HB PECAS' in t or 'PECAS' in t: return 'H.B.PECAS'
    if 'ESTOQUE PECAS' in t or 'EST.PECAS' in t: return 'ESTOQUE PECAS'
    if 'OBRIG' in t: return 'OBRIG. A PAGA'
    if 'ADIANTAMENTO' in t: return 'ADIANTAMENTOS'
    if 'TRANSITORIA' in t: return 'TRANSITORIA'
    if 'DIF_TRANS_ADIANT' in t: return 'DIF_TRANS_ADIANT'
    return 'OUTROS'

ITENS = [
    ('CARTEIRA', 'CARTEIRA'), ('MERCADO PAGO', 'MERCADO PAGO'),('VEICULO', 'VEICULO'), ('SEGURADORA', 'SEGURADORA'),
    ('GARANTIA', 'GARANTIA'), ('BANCOS', 'BANCOS'), ('CARTOES', 'CARTOES'),
    ('NOVOS PAGOS', 'NOVOS PAGOS'),('USADOS PAGOS', 'USADOS PAGOS'),('FUNDAO NOVOS', 'FUNDAO NOVOS'),
    ('FIDIC', 'FIDIC'), ('H.B.PECAS', 'H.B.PECAS'),
    ('ESTOQUE PECAS', 'ESTOQUE PECAS'), ('OBRIG. A PAGA', 'OBRIG. A PAGA'),('ADIANTAMENTOS', 'ADIANTAMENTOS'),
    ('TRANSITORIA', 'TRANSITORIA'), ('DIF_TRANS_ADIANT', 'DIF_TRANS_ADIANT')
]

ITENS_MANUAIS = [
    ('NOVOS PAGOS', 'NOVOS PAGOS'),
    ('USADOS PAGOS', 'USADOS PAGOS'),
    ('FUNDAO NOVOS', 'FUNDAO NOVOS'),
    ('FIDIC', 'FIDIC'), ('H.B.PECAS', 'H.B.PECAS'),
    ('ESTOQUE PECAS', 'ESTOQUE PECAS')
]
# Carrega Valores manuais do banco
def carregar_valores_manuais_do_banco(data_ref):
    # Garante que data_ref é objeto date
    if isinstance(data_ref, str):
        data_ref = date.fromisoformat(data_ref)
        
    db = SessionLocal()
    valores = {
        'MATRIZ': {k: '0,00' for k,_ in ITENS_MANUAIS},
        'WS': {k: '0,00' for k,_ in ITENS_MANUAIS},
        'EUSEBIO': {k: '0,00' for k,_ in ITENS_MANUAIS}
    }
    valores_qtd = {
        'MATRIZ': {'NOVOS PAGOS': 0, 'USADOS PAGOS': 0},
        'WS': {'NOVOS PAGOS': 0, 'USADOS PAGOS': 0},
        'EUSEBIO': {'NOVOS PAGOS': 0, 'USADOS PAGOS': 0}
    }
    registros = db.query(PosicaoDiaria).filter(
        PosicaoDiaria.data == data_ref, # <- agora é date
        PosicaoDiaria.tipo_titulo.in_([k for k,_ in ITENS_MANUAIS])
    ).all()
    for reg in registros:
        valores[reg.empresa][reg.tipo_titulo] = formatar_br(reg.valor).replace('R$ ', '')
        if reg.tipo_titulo in ['NOVOS PAGOS', 'USADOS PAGOS']:
            valores_qtd[reg.empresa][reg.tipo_titulo] = int(reg.qtd_veiculos or 0)
    db.close()
    return valores, valores_qtd
    
# Salvar posicao no banco de dados 
def salvar_posicao_no_banco(df, data_ref, modo='novo'):
    db = SessionLocal()
    usuario_logado = st.session_state.get('email', 'sistema')

    df = df.copy()

    # 1. GARANTE QUE TODAS AS COLUNAS EXISTEM
    if 'Qtd' not in df.columns: df['Qtd'] = 0
    if 'ValorMedio' not in df.columns: df['ValorMedio'] = 0.0

    # 2. LIMPEZA EXTREMA
    df['Saldo'] = pd.to_numeric(df['Saldo'], errors='coerce').fillna(0.0)
    df['Qtd'] = pd.to_numeric(df['Qtd'], errors='coerce').fillna(0).astype(int)
    df['ValorMedio'] = pd.to_numeric(df['ValorMedio'], errors='coerce').fillna(0.0)
    df['ValorMedio'] = df['ValorMedio'].replace([np.inf, -np.inf], 0.0)

    df['Empresa'] = df['Empresa'].astype(str).str.strip()
    df['Tipo de Título'] = df['Tipo de Título'].astype(str).str.strip()

    # 3. TIRA LINHAS LIXO
    df = df[(df['Empresa']!= '') & (df['Empresa']!= 'nan') & (df['Empresa']!= 'None')]
    df = df[(df['Tipo de Título']!= '') & (df['Tipo de Título']!= 'nan') & (df['Tipo de Título']!= 'None')]

    if isinstance(data_ref, str):
        data_ref = date.fromisoformat(data_ref)

    if df.empty:
        st.warning("Nenhuma linha válida para salvar")
        db.close()
        return

    try:
        if modo == 'manutencao':
            for _, row in df.iterrows():
                reg = db.query(PosicaoDiaria).filter(
                    PosicaoDiaria.data == data_ref,
                    PosicaoDiaria.empresa == row['Empresa'],
                    PosicaoDiaria.tipo_titulo == row['Tipo de Título']
                ).first()
                if reg:
                    reg.valor = float(row['Saldo'])
                    reg.qtd_veiculos = int(row['Qtd'])
                    reg.valor_medio = float(row['ValorMedio'])
                    reg.criado_por = usuario_logado
                else:
                    db.add(PosicaoDiaria(
                        data=data_ref, empresa=row['Empresa'], tipo_titulo=row['Tipo de Título'],
                        valor=float(row['Saldo']), qtd_veiculos=int(row['Qtd']),
                        valor_medio=float(row['ValorMedio']), criado_por=usuario_logado
                    ))
        else:
            db.query(PosicaoDiaria).filter(PosicaoDiaria.data == data_ref).delete()
            objs = []
            for _, row in df.iterrows():
                objs.append(PosicaoDiaria(
                    data=data_ref, empresa=row['Empresa'], tipo_titulo=row['Tipo de Título'],
                    valor=float(row['Saldo']), qtd_veiculos=int(row['Qtd']),
                    valor_medio=float(row['ValorMedio']), criado_por=usuario_logado
                ))
            db.add_all(objs)

        db.commit()
        
        # <- ADICIONA ESSA LINHA AQUI: Limpa o cache pra forçar recalcular os cards
        st.cache_data.clear() 
        
        st.success(f"{len(df)} registros salvos com sucesso!")

    except Exception as e:
        db.rollback()
        st.error(f"Erro ao salvar: {e}")
        st.write("DataFrame que deu erro:")
        st.dataframe(df)
    finally:
        db.close()
    
def zerar_banco():
    db = SessionLocal()
    qtd = db.query(PosicaoDiaria).count()
    db.query(PosicaoDiaria).delete()
    db.commit()
    db.close()
    return qtd
    
# Gerar excel exporta empresas selecionadas.
def gerar_excel(df_para_exportar, empresas_selecionadas, data_titulo_str):
    output = BytesIO()
    data_excel = date.fromisoformat(data_titulo_str).strftime('%d/%m/%Y')
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet('POSICAO DIARIA')
        writer.sheets['POSICAO DIARIA'] = worksheet
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        border_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold = Font(bold=True, size=11); center = Alignment(horizontal='center', vertical='center'); right = Alignment(horizontal='right', vertical='center')
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        cell_titulo = worksheet.cell(row=1, column=1, value="POSIÇÃO FINANCEIRA DIÁRIA"); cell_titulo.font = Font(bold=True, size=14); cell_titulo.alignment = center; cell_titulo.border = border_fina
        worksheet.cell(row=1, column=7, value='DATA').font = bold; worksheet.cell(row=1, column=7).alignment = center; worksheet.cell(row=1, column=7).border = border_fina
        worksheet.cell(row=1, column=8, value=data_excel).border = border_fina

        col_inicio = 1
        for emp in ['MATRIZ', 'WS', 'EUSEBIO']:
            if emp not in empresas_selecionadas: col_inicio += 3; continue
            worksheet.merge_cells(start_row=2, start_column=col_inicio, end_row=2, end_column=col_inicio+1)
            cell_emp = worksheet.cell(row=2, column=col_inicio, value=emp); cell_emp.font = bold; cell_emp.alignment = center; cell_emp.fill = header_fill; cell_emp.border = border_fina
            worksheet.cell(row=2, column=col_inicio+1).fill = header_fill; worksheet.cell(row=2, column=col_inicio+1).border = border_fina
            col_inicio += 3

        linha_temp = 3; col_inicio = 1
        for emp in ['MATRIZ', 'WS', 'EUSEBIO']:
            if emp not in empresas_selecionadas: col_inicio += 3; continue
            worksheet.column_dimensions[get_column_letter(col_inicio)].width = 25
            worksheet.column_dimensions[get_column_letter(col_inicio+1)].width = 18
            worksheet.cell(row=linha_temp, column=col_inicio, value='DESCRICAO').font = bold; worksheet.cell(row=linha_temp, column=col_inicio).fill = header_fill; worksheet.cell(row=linha_temp, column=col_inicio).border = border_fina
            worksheet.cell(row=linha_temp, column=col_inicio+1, value='VALORES').font = bold; worksheet.cell(row=linha_temp, column=col_inicio+1).fill = header_fill; worksheet.cell(row=linha_temp, column=col_inicio+1).border = border_fina
            col_inicio += 3

        linha_dados = 4
        for item_chave, item_nome in ITENS:
            col_inicio = 1
            for emp in ['MATRIZ', 'WS', 'EUSEBIO']:
                if emp not in empresas_selecionadas: col_inicio += 3; continue
                total_valor = df_para_exportar[(df_para_exportar['Tipo de Título'] == item_chave) & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                total_qtd = df_para_exportar[(df_para_exportar['Tipo de Título'] == item_chave) & (df_para_exportar['Empresa'] == emp)]['Qtd'].sum()
                if item_chave == 'DIF_TRANS_ADIANT':
                    trans_valor = df_para_exportar[(df_para_exportar['Tipo de Título'] == 'TRANSITORIA') & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                    adiant_valor = df_para_exportar[(df_para_exportar['Tipo de Título'] == 'ADIANTAMENTOS') & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                    total_valor = adiant_valor - trans_valor if trans_valor > 0 else 0.0
                cell_desc = worksheet.cell(row=linha_dados, column=col_inicio, value=item_nome); cell_desc.border = border_fina
                if item_chave in ['NOVOS PAGOS', 'USADOS PAGOS']:
                    valor_formatado = formatar_br(total_valor).replace('R$ ', '') # <- CORRIGIDO
                    cell_valor = worksheet.cell(row=linha_dados, column=col_inicio+1, value=f"{valor_formatado} - {int(total_qtd)}")
                    cell_valor.alignment = right; cell_valor.border = border_fina
                else:
                    cell_valor = worksheet.cell(row=linha_dados, column=col_inicio+1, value=total_valor); cell_valor.alignment = right; cell_valor.number_format = 'R$ #,##0.00'; cell_valor.border = border_fina
                col_inicio += 3
            linha_dados += 1

        col_inicio = 1
        for emp in ['MATRIZ', 'WS', 'EUSEBIO']:
            if emp not in empresas_selecionadas: col_inicio += 3; continue
            total_geral = 0.0
            for item_chave, item_nome in ITENS:
                total = df_para_exportar[(df_para_exportar['Tipo de Título'] == item_chave) & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                if item_chave == 'DIF_TRANS_ADIANT':
                    trans_valor = df_para_exportar[(df_para_exportar['Tipo de Título'] == 'TRANSITORIA') & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                    adiant_valor = df_para_exportar[(df_para_exportar['Tipo de Título'] == 'ADIANTAMENTOS') & (df_para_exportar['Empresa'] == emp)]['Saldo'].sum()
                    total = adiant_valor - trans_valor if trans_valor > 0 else 0.0
                if item_chave == 'OBRIG. A PAGA':
                    total_geral -= total
                elif item_chave not in ['TRANSITORIA', 'DIF_TRANS_ADIANT']:
                    total_geral += total
            worksheet.cell(row=linha_dados, column=col_inicio, value='TOTAL').font = bold; worksheet.cell(row=linha_dados, column=col_inicio).border = border_fina
            cell_total = worksheet.cell(row=linha_dados, column=col_inicio+1, value=total_geral); cell_total.font = bold; cell_total.alignment = right; cell_total.number_format = 'R$ #,##0.00'; cell_total.border = border_fina
            col_inicio += 3
    return output.getvalue()
# cards
with tab1:
    with st.sidebar:
        st.markdown("### Data de Referência do Lançamento")
        DATA_REF_DATE = st.date_input("Selecione a Data", value=date.today(), format="DD/MM/YYYY", key="data_lancamento")

        st.markdown("### Filtros")
        empresas_selecionadas = st.multiselect("Empresas", ['MATRIZ', 'WS', 'EUSEBIO'], default=['MATRIZ', 'WS', 'EUSEBIO'])
        st.divider()
        st.markdown("### Exportar")
        st.markdown("### Resumo do Dia")
        col_m, col_ws, col_e = st.columns(3)
        empresas_cards = {'MATRIZ': col_m, 'WS': col_ws, 'EUSEBIO': col_e}
        
        from sqlalchemy import func # <- COLOQUEI O IMPORT AQUI DENTRO
        
        # CACHE CURTO PRA FORÇAR ATUALIZAR
        @st.cache_data(ttl=10)
        def get_total_empresa(data, empresa):
            db = SessionLocal()
            try:
                total = db.query(func.sum(PosicaoDiaria.valor)).filter(
                    PosicaoDiaria.data == data,
                    PosicaoDiaria.empresa == empresa
                ).scalar()
                return float(total or 0.0)
            finally:
                db.close()
        
        @st.cache_data(ttl=10)
        def get_variacao_empresa(data_atual, data_anterior, empresa):
            if not data_anterior: return 0.0
            total_atual = get_total_empresa(data_atual, empresa)
            total_anterior = get_total_empresa(data_anterior, empresa)
            if total_anterior == 0: return 0.0
            return ((total_atual - total_anterior) / total_anterior) * 100
        
        
        for emp, col in empresas_cards.items():
            if emp not in empresas_selecionadas: continue
            with col:
                total_hoje = get_total_empresa(DATA_REF_DATE, emp)
                data_ontem_date = DATA_REF_DATE - timedelta(days=1)
                variacao = get_variacao_empresa(DATA_REF_DATE, data_ontem_date, emp)
        
                # SETA PRA CIMA OU PRA BAIXO
                delta_cor = "normal"
                if variacao < 0: delta_cor = "inverse"
        
                st.metric(
                    label=f"TOTAL {emp}",
                    value=formatar_br(total_hoje),
                    delta=f"{variacao:.2f}%" if variacao is not None else None,
                    delta_color=delta_cor
                )
        st.divider()
        
        with st.expander("📝 Lançamento Manual / Manutenção - Editável", expanded=True):
            col_cal, col_btn = st.columns([3,1])
            with col_cal:
                DATA_MANUTENCAO_DATE = st.date_input("📅 Selecione a Data para Manutenção", value=DATA_REF_DATE, format="DD/MM/YYYY", key="data_manutencao")
                DATA_MANUTENCAO = DATA_MANUTENCAO_DATE.strftime("%Y-%m-%d")
            with col_btn:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("📂 Carregar Dados da Data", key="btn_carregar_manut"):
                    db = SessionLocal()
                    registros = db.query(PosicaoDiaria).filter(PosicaoDiaria.data == DATA_MANUTENCAO_DATE).all()
                    db.close()
                    if registros:
                        dados = [{'Tipo de Título': r.tipo_titulo, 'Empresa': r.empresa, 'Saldo': r.valor, 'Qtd': r.qtd_veiculos, 'ValorMedio': r.valor_medio} for r in registros]
                        st.session_state['df_carregado_manut'] = pd.DataFrame(dados)
                        st.success(f"Dados de {DATA_MANUTENCAO_DATE.strftime('%d/%m/%Y')} carregados.")
                    else:
                        st.warning("Nenhum dado encontrado para esta data.")
                        st.session_state['df_carregado_manut'] = pd.DataFrame()
                    st.rerun()
        
            valores_iniciais, valores_qtd_iniciais = carregar_valores_manuais_do_banco(DATA_MANUTENCAO)
        
            
        if 'df_carregado_manut' in st.session_state and not st.session_state['df_carregado_manut'].empty:
            df_temp = st.session_state['df_carregado_manut']
            for emp in ['MATRIZ', 'WS', 'EUSEBIO']:
                for item_chave, _ in ITENS_MANUAIS:
                    val = df_temp[(df_temp['Empresa']==emp) & (df_temp['Tipo de Título']==item_chave)]['Saldo'].sum()
                    qtd = df_temp[(df_temp['Empresa']==emp) & (df_temp['Tipo de Título']==item_chave)]['Qtd'].sum()
                    valores_iniciais[emp][item_chave] = formatar_br(val).replace('R$ ', '')
                    if item_chave in ['NOVOS PAGOS', 'USADOS PAGOS']:
                        valores_qtd_iniciais[emp][item_chave] = int(qtd)

        uploaded_files = st.file_uploader("📁 Arraste os 4 arquivos RFN aqui - Só se for lançamento novo", type=['xlsx', 'xls'], accept_multiple_files=True, key="uploader_tab1")
        dfs = {}
        if uploaded_files:
            dfs = {file.name: pd.read_excel(file) for file in uploaded_files}
        
        # Carrega arquivos para possição analitica
        def carregar_posicao_analitica():
            if not uploaded_files: return pd.DataFrame()
        
            # PEGA PELO RFN003 + RECEBER
            df_raw = None
            for nome, df_temp in dfs.items():
                if 'RFN003' in nome.upper() and 'RECEBER' in nome.upper():
                    df_raw = df_temp
                    break
        
            if df_raw is None: return pd.DataFrame()
            df_raw.columns = df_raw.columns.str.strip()
            dados = []
            for empresa in df_raw['Empresa'].dropna().unique():
                df_empresa = df_raw[df_raw['Empresa'] == empresa]
                empresa_norm = detectar_empresa(empresa)
                for _, row in df_empresa.iterrows():
                    try:
                        agente = str(row['Agente Cobrador']).strip()
                        titulo = str(row['Nro Titulo']).strip()
                        saldo = converter_valor_br(row['Saldo'])
                        if titulo == '' or titulo.upper() == 'NAN' or saldo <= 0: continue
                        tipo = normalizar_tipo(agente)
                        dados.append({'Tipo de Título': tipo, 'Empresa': empresa_norm, 'Saldo': saldo})
                    except: continue
            return pd.DataFrame(dados) if dados else pd.DataFrame()
                
        def carregar_obrigacoes():
            if not uploaded_files: return pd.DataFrame()
        
            df_obrig_raw = None
            for nome, df_temp in dfs.items():
                if 'RFN003' in nome.upper() and 'PAGAR' in nome.upper():
                    df_obrig_raw = df_temp
                    break
        
            if df_obrig_raw is None: return pd.DataFrame()
            obrig_dict = {'MATRIZ': 0.0, 'WS': 0.0, 'EUSEBIO': 0.0}
            for i in range(len(df_obrig_raw)):
                linha_toda = " ".join([str(x) for x in df_obrig_raw.iloc[i] if pd.notna(x)]).upper()
                if 'RESUMO' in linha_toda and 'EMPRESA' in linha_toda:
                    for j in range(1, 6):
                        if i+j >= len(df_obrig_raw): break
                        emp = str(df_obrig_raw.iloc[i + j, 0])
                        val = converter_valor_br(df_obrig_raw.iloc[i + j, 9])
                        if 'EUSEBIO' in emp.upper(): obrig_dict['EUSEBIO'] = val
                        elif 'MATRIZ' in emp.upper(): obrig_dict['MATRIZ'] = val
                        elif 'WS' in emp.upper(): obrig_dict['WS'] = val
                    break
            linhas = [{'Tipo de Título': 'OBRIG. A PAGA', 'Empresa': e, 'Saldo': v} for e,v in obrig_dict.items() if v > 0]
            return pd.DataFrame(linhas)
                
        def carregar_creditos_nao_identificados():
            if not uploaded_files: return pd.DataFrame()
        
            df_cred_raw = None
            for nome, df_temp in dfs.items():
                if 'RFN024' in nome.upper():
                    df_cred_raw = df_temp
                    break
        
            if df_cred_raw is None: return pd.DataFrame()
            credito_dict = {'MATRIZ': 0.0, 'WS': 0.0, 'EUSEBIO': 0.0}
            for i in range(len(df_cred_raw)):
                if not str(df_cred_raw.iloc[i, 0]).isdigit(): continue
                emp = detectar_empresa(str(df_cred_raw.iloc[i, 2]))
                sal = converter_valor_br(df_cred_raw.iloc[i, df_cred_raw.shape[1]-1])
                if emp in credito_dict and sal > 0: credito_dict[emp] += sal
            linhas = [{'Tipo de Título': 'TRANSITORIA', 'Empresa': e, 'Saldo': v} for e,v in credito_dict.items() if v > 0]
            return pd.DataFrame(linhas)
                
        def carregar_adiantamentos():
            if not uploaded_files: return pd.DataFrame()
        
            df_ad_raw = None
            for nome, df_temp in dfs.items():
                if 'RFN013' in nome.upper():
                    df_ad_raw = df_temp
                    break
        
            if df_ad_raw is None: return pd.DataFrame()
            adiant_dict = {'MATRIZ': 0.0, 'WS': 0.0, 'EUSEBIO': 0.0}
            for i in range(1, len(df_ad_raw)):
                emp = detectar_empresa(str(df_ad_raw.iloc[i, 3]))
                sal = converter_valor_br(str(df_ad_raw.iloc[i, 6]))
                if emp in adiant_dict: adiant_dict[emp] += sal
            linhas = [{'Tipo de Título': 'ADIANTAMENTOS', 'Empresa': e, 'Saldo': v} for e,v in adiant_dict.items() if v > 0]
            return pd.DataFrame(linhas)
        
        # def carregar_manuais_adiantamentos()
        def carregar_manuais(valores_digitados, valores_qtd_digitados):
            dados = []
            for empresa, itens in valores_digitados.items():
                for tipo, valor_str in itens.items():
                    valor = converter_valor_br(valor_str)
                    qtd = int(valores_qtd_digitados[empresa].get(tipo, 0)) # força int
                    valor_medio = float(valor / qtd) if qtd > 0 else 0.0 # força float
                    dados.append({
                        'Tipo de Título': tipo,
                        'Empresa': empresa,
                        'Saldo': float(valor),
                        'Qtd': qtd,
                        'ValorMedio': valor_medio
                    })
            return pd.DataFrame(dados)
        
        col_m, col_ws, col_e = st.columns(3)
        valores_digitados = {'MATRIZ': {}, 'WS': {}, 'EUSEBIO': {}}
        valores_qtd_digitados = {'MATRIZ': {}, 'WS': {}, 'EUSEBIO': {}}
        empresas_col = {'MATRIZ': col_m, 'WS': col_ws, 'EUSEBIO': col_e}
        
        for emp, col in empresas_col.items():
            with col:
                st.markdown(f"**{emp}**")
                for item_chave, item_nome in ITENS_MANUAIS:
                    st.markdown(f"{item_nome}")
                    key_valor = f"{emp}_{item_chave}_edit_{DATA_MANUTENCAO.replace('-', '')}"
                    key_qtd = f"{emp}_QTD_{item_chave}_edit_{DATA_MANUTENCAO.replace('-', '')}"
        
                    valores_digitados[emp][item_chave] = st.text_input(
                        label="",
                        value=valores_iniciais[emp][item_chave],
                        key=key_valor,
                        label_visibility="collapsed"
                    )
                    if item_chave in ['NOVOS PAGOS', 'USADOS PAGOS']:
                        valores_qtd_digitados[emp][item_chave] = st.number_input(
                            "Qtd",
                            value=int(valores_qtd_iniciais[emp][item_chave]), # <- INT
                            key=key_qtd,
                            min_value=0,
                            step=1 # <- STEP
                        )
        # Botao salvar e atualizar
        if st.button(f"💾 Salvar / Atualizar {DATA_MANUTENCAO_DATE.strftime('%d/%m/%Y')}", key="btn_salvar_manut"):
            if uploaded_files:
                lista_df = [carregar_posicao_analitica(), carregar_obrigacoes(), carregar_creditos_nao_identificados(), carregar_adiantamentos(), carregar_manuais(valores_digitados, valores_qtd_digitados)]
                lista_df = [df for df in lista_df if not df.empty]
                if lista_df:
                    df = pd.concat(lista_df, ignore_index=True)
                    
                    # BLINDAGEM TOTAL - CRIA AS 3 COLUNAS SEMPRE
                    df['Saldo'] = pd.to_numeric(df['Saldo'], errors='coerce').fillna(0.0)
                    df['Qtd'] = pd.to_numeric(df.get('Qtd', 0), errors='coerce').fillna(0).astype(int)
                    df['ValorMedio'] = pd.to_numeric(df.get('ValorMedio', 0.0), errors='coerce').fillna(0.0)
                    df['ValorMedio'] = df['ValorMedio'].replace([np.inf, -np.inf], 0.0)
                    
                    empresas = df['Empresa'].unique()
                    novas_linhas = []
                    for emp in empresas:
                        trans = df[(df['Empresa'] == emp) & (df['Tipo de Título'] == 'TRANSITORIA')]['Saldo'].sum()
                        adiant = df[(df['Empresa'] == emp) & (df['Tipo de Título'] == 'ADIANTAMENTOS')]['Saldo'].sum()
                        if trans > 0:
                            dif = adiant - trans
                            if dif!= 0: novas_linhas.append({'Tipo de Título': 'DIF_TRANS_ADIANT', 'Empresa': emp, 'Saldo': float(dif), 'Qtd': 0, 'ValorMedio': 0.0})
                    if novas_linhas: df = pd.concat([df, pd.DataFrame(novas_linhas)], ignore_index=True)
                    
                    salvar_posicao_no_banco(df, DATA_MANUTENCAO_DATE, modo='novo')
                    st.session_state['df_final'] = df
                    st.success(f"Dados de {DATA_MANUTENCAO_DATE.strftime('%d/%m/%Y')} ATUALIZADOS no banco!")
            else:
                df_manual = carregar_manuais(valores_digitados, valores_qtd_digitados)
                salvar_posicao_no_banco(df_manual, DATA_MANUTENCAO_DATE, modo='manutencao')
                if 'df_carregado_manut' in st.session_state:
                    del st.session_state['df_carregado_manut']
                st.success(f"Manutenção salva! Data: {DATA_MANUTENCAO_DATE.strftime('%d/%m/%Y')}")
            st.rerun()
        
        if 'df_final' in st.session_state:
            df = st.session_state['df_final']
            with st.sidebar:
                excel_data = gerar_excel(df, empresas_selecionadas, DATA_REF_DATE.strftime("%Y-%m-%d"))
                st.download_button(label="📊 Gerar Planilha Única", data=excel_data, file_name=f"Posicao_Financeira_{DATA_REF_DATE.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            col1, col2, col3 = st.columns(3)
            empresas_cols = {'MATRIZ': col1, 'WS': col2, 'EUSEBIO': col3}
            for emp, col in empresas_cols.items():
                if emp not in empresas_selecionadas: continue
                with col:
                    dados_tabela = []
                    total_geral = 0.0
                    for item_chave, item_nome in ITENS:
                        total = df[(df['Tipo de Título'] == item_chave) & (df['Empresa'] == emp)]['Saldo'].sum()
                        total_qtd = df[(df['Tipo de Título'] == item_chave) & (df['Empresa'] == emp)]['Qtd'].sum()
                        if item_chave == 'DIF_TRANS_ADIANT':
                            trans_valor = df[(df['Tipo de Título'] == 'TRANSITORIA') & (df['Empresa'] == emp)]['Saldo'].sum()
                            adiant_valor = df[(df['Tipo de Título'] == 'ADIANTAMENTOS') & (df['Empresa'] == emp)]['Saldo'].sum()
                            total = adiant_valor - trans_valor if trans_valor > 0 else 0.0
                        if item_chave == 'OBRIG. A PAGA':
                            total_geral -= total
                        elif item_chave not in ['TRANSITORIA', 'DIF_TRANS_ADIANT']:
                            total_geral += total
                        if item_chave in ['NOVOS PAGOS', 'USADOS PAGOS']:
                            valor_formatado = formatar_br(total).replace('R$ ', '')
                            dados_tabela.append({"DESCRICAO": item_nome, "VALORES": f"{valor_formatado} - {int(total_qtd)}"})
                        else:
                            dados_tabela.append({"DESCRICAO": item_nome, "VALORES": formatar_br(total)})
                    dados_tabela.append({"DESCRICAO": "TOTAL", "VALORES": formatar_br(total_geral)})
                    df_mostrar = pd.DataFrame(dados_tabela)
                    st.dataframe(df_mostrar, hide_index=True, use_container_width=True, height=680, column_config={
                        "DESCRICAO": st.column_config.TextColumn("DESCRICAO"),
                        "VALORES": st.column_config.TextColumn("VALORES")
                    })
        # ---- Tabela2 -------
        with tab2:
            st.header("📊 Relatórios e Auditoria")
            
            with st.expander("⚠️ Zona Perigosa - Apagar Dados"):
                st.warning("Isso vai apagar TODAS as datas do banco. Use só pra zerar e recomeçar.")
                if st.button("🗑️ ZERAR BANCO COMPLETO"):
                    qtd_apagada = zerar_banco()
                    st.success(f"Banco zerado! {qtd_apagada} registros apagados.")
                    st.rerun()
        
            st.divider()
            st.subheader("Gerar Relatório de Auditoria por Período")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                data_inicio = st.date_input("Data Inicial", value=date.today() - timedelta(days=7))
            with col2:
                data_fim = st.date_input("Data Final", value=date.today())
            with col3:
                db = SessionLocal()
                usuarios = [u[0] for u in db.query(Usuarios.email).filter(Usuarios.ativo==True).all()]
                db.close()
                usuario_filtro = st.selectbox("Filtrar por Usuário", ["Todos"] + usuarios)
            
            if st.button("📄 Gerar Relatório PDF", type="primary"):
                db = SessionLocal()
                query = db.query(PosicaoDiaria).filter(
                    PosicaoDiaria.data >= data_inicio, # <- manda date direto
                    PosicaoDiaria.data <= data_fim     # <- manda date direto
                )
                if usuario_filtro!= "Todos":
                    query = query.filter(PosicaoDiaria.criado_por == usuario_filtro)
                registros = query.all()
                db.close()
        
                if not registros:
                    st.warning("Nenhum dado encontrado no período.")
                else:
                    df_rel = pd.DataFrame([{
                        'Data': r.data, 'Empresa': r.empresa, 'Item': r.tipo_titulo, 
                        'Valor': r.valor, 'Qtd': r.qtd_veiculos, 'Lançado por': r.criado_por
                    } for r in registros])
        
                    # Gera PDF
                    from reportlab.pdfgen import canvas
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib.units import cm
                    from reportlab.platypus import Table, TableStyle
                    from reportlab.lib import colors
        
                    buffer = BytesIO()
                    c = canvas.Canvas(buffer, pagesize=A4)
                    width, height = A4
        
                    c.setFont("Helvetica-Bold", 16)
                    c.drawString(2*cm, height - 2*cm, "RELATÓRIO DE AUDITORIA - POSIÇÃO FINANCEIRA")
                    c.setFont("Helvetica", 10)
                    c.drawString(2*cm, height - 2.8*cm, f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}")
                    c.drawString(2*cm, height - 3.3*cm, f"Usuário Filtro: {usuario_filtro}")
        
                    # Resumo por empresa
                    resumo = df_rel.groupby('Empresa')['Valor'].sum().reset_index()
                    data_resumo = [['Empresa', 'Total Lançado']] + [[row['Empresa'], formatar_br(row['Valor'])] for _, row in resumo.iterrows()]
                    t_resumo = Table(data_resumo, colWidths=[6*cm, 4*cm])
                    t_resumo.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, colors.black)]))
                    t_resumo.wrapOn(c, width, height)
                    t_resumo.drawOn(c, 2*cm, height - 6*cm)
        
                    c.showPage()
                    c.save()
                    pdf_data = buffer.getvalue()
        
                    st.success(f"{len(df_rel)} registros encontrados.")
                    st.dataframe(df_rel, use_container_width=True, hide_index=True)
        
                    st.download_button(
                        label="⬇️ Baixar Relatório PDF",
                        data=pdf_data,
                        file_name=f"Relatorio_Auditoria_{data_inicio}_{data_fim}.pdf",
                        mime="application/pdf"
                    )
